"""Export DataFrames to downloadable bytes (CSV or Excel)."""

from __future__ import annotations

import re
from enum import Enum
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .constants import (
    CSV_ENCODING,
    GROUP_SEPARATOR_AFTER_HEADERS,
    GROUP_SEPARATOR_COLUMN_WIDTH,
    GROUP_SEPARATOR_FILL_HEX,
    HEADER_COLUMN_WIDTH_PX,
    HEADER_FILL_HEX_BY_COLUMN,
    HEADER_ROW_HEIGHT_PX,
    LONG_NUMERIC_ID_COLUMNS,
)
from .id_format import coerce_dataframe_long_ids, plain_id_string

# Invisible LTR mark — Excel keeps the cell as text; digits stay visible without E+ notation.
_XLSX_TEXT_PREFIX = "\u200e"


class ExportFormat(str, Enum):
    CSV = "csv"
    XLSX = "xlsx"


def _export_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Replace nulls with blank strings so every column shows empty cells, not NaN."""
    return df.fillna("")


def _column_index_for_header(ws, header: str) -> int | None:
    """1-based column index where row 1 equals ``header`` (first match)."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v == header:
            return c
    return None


def _column_indices_for_header(ws, header: str) -> list[int]:
    """1-based column indices where row 1 equals ``header`` (all matches — duplicate headers)."""
    return [
        c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value == header
    ]


def _sheet_bottom_row(ws) -> int:
    """Last row index that contains data (openpyxl ``max_row`` plus dimension fallback)."""
    last = ws.max_row or 1
    dim = ws.calculate_dimension()
    if dim and ":" in dim:
        corner = dim.split(":")[-1].strip()
        m = re.search(r"(\d+)$", corner)
        if m:
            last = max(last, int(m.group(1)))
    return last


def _insert_group_separator_columns(workbook) -> set[int]:
    """
    Insert narrow navy columns between header groups. Inserts from right to left.
    Returns final 1-based indices of separator columns (for skipping header styling).
    """
    ws = workbook["Sheet1"]
    insert_before: list[int] = []
    for anchor in GROUP_SEPARATOR_AFTER_HEADERS:
        idx = _column_index_for_header(ws, anchor)
        if idx is not None:
            insert_before.append(idx + 1)

    sep_fill = PatternFill(patternType="solid", fgColor=GROUP_SEPARATOR_FILL_HEX)
    separator_final_cols: list[int] = []

    for pos in sorted(set(insert_before), reverse=True):
        ws.insert_cols(pos, amount=1)
        separator_final_cols = [c + 1 if c >= pos else c for c in separator_final_cols]
        separator_final_cols.append(pos)

        letter = get_column_letter(pos)
        ws.column_dimensions[letter].width = GROUP_SEPARATOR_COLUMN_WIDTH
        bottom = _sheet_bottom_row(ws)
        for row_idx in range(1, bottom + 1):
            cell = ws.cell(row_idx, pos)
            cell.value = None
            cell.fill = sep_fill

    return set(separator_final_cols)


def _enforce_separator_widths(ws, separator_cols: set[int]) -> None:
    """Re-apply narrow width (pandas / merges can reset column_dimensions)."""
    for c in separator_cols:
        ws.column_dimensions[get_column_letter(c)].width = GROUP_SEPARATOR_COLUMN_WIDTH


def _apply_header_dimensions(ws, separator_cols: set[int]) -> None:
    """
    Size the header row to match the template (87×146 px at 96dpi).

    Row height is set in **points** (openpyxl); column width in Excel **character units**
    (approximate mapping from pixels for default-font columns). Separator columns stay narrow.
    """
    # Pixels → row height in points (Excel UI px @ 96dpi).
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT_PX * 72.0 / 96.0

    # Pixels → Excel column width (common approximation for Calibri-scale columns).
    col_w = max(8.43, (HEADER_COLUMN_WIDTH_PX - 5.0) / 7.0)
    for c in range(1, ws.max_column + 1):
        if c in separator_cols:
            continue
        ws.column_dimensions[get_column_letter(c)].width = col_w


def _repaint_separator_columns_to_last_row(ws, separator_cols: set[int]) -> None:
    """Navy fill for every row through the sheet's last used row (styling can strip fills)."""
    sep_fill = PatternFill(patternType="solid", fgColor=GROUP_SEPARATOR_FILL_HEX)
    last = _sheet_bottom_row(ws)
    for c in separator_cols:
        for r in range(1, last + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = sep_fill


def _rewrite_openpyxl_long_id_cells(workbook) -> None:
    """
    Open Excel sheet and force ID columns to real text cells (resolve columns by header text).
    """
    ws = workbook["Sheet1"]

    for col_name in LONG_NUMERIC_ID_COLUMNS:
        col_i = _column_index_for_header(ws, col_name)
        if col_i is None:
            continue
        letter = get_column_letter(col_i)
        ws[f"{letter}1"].number_format = "@"

        bottom = _sheet_bottom_row(ws)
        for row in range(2, bottom + 1):
            cell = ws[f"{letter}{row}"]
            raw = cell.value
            if raw is None or raw == "":
                cell.value = ""
            else:
                text = plain_id_string(raw)
                cell.value = _XLSX_TEXT_PREFIX + text if text else ""
            cell.number_format = "@"


def _pattern_fill(hex_rgb: str) -> PatternFill:
    """New PatternFill per application — avoid sharing mutable styles across cells/workbooks."""
    return PatternFill(patternType="solid", fgColor=hex_rgb)


_HEADER_WRAP_ALIGN = Alignment(
    wrap_text=True,
    vertical="center",
    horizontal="center",
)
_HEADER_FONT_BOLD = Font(bold=True)


def _apply_wrap_text_alignment(workbook, *, skip_columns: set[int]) -> None:
    """Bold header row and Wrap Text; skip narrow navy separator columns."""
    ws = workbook["Sheet1"]
    for row in ws.iter_rows(
        min_row=1,
        max_row=1,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            if cell.column in skip_columns:
                continue
            cell.alignment = _HEADER_WRAP_ALIGN
            cell.font = _HEADER_FONT_BOLD


def _apply_header_row_fills(workbook) -> None:
    """Apply template header colors (resolve columns by header text)."""
    ws = workbook["Sheet1"]
    for col_name, hex_rgb in HEADER_FILL_HEX_BY_COLUMN.items():
        for col_i in _column_indices_for_header(ws, col_name):
            letter = get_column_letter(col_i)
            ws[f"{letter}1"].fill = _pattern_fill(hex_rgb)


def _finalize_xlsx_bytes(raw: BytesIO) -> bytes:
    """
    Reload and re-save with openpyxl: group separators, IDs, fills, header formatting.
    """
    raw.seek(0)
    wb = load_workbook(raw)
    sep_cols = _insert_group_separator_columns(wb)
    _rewrite_openpyxl_long_id_cells(wb)
    _apply_header_row_fills(wb)
    _apply_wrap_text_alignment(wb, skip_columns=sep_cols)
    _repaint_separator_columns_to_last_row(wb["Sheet1"], sep_cols)
    _apply_header_dimensions(wb["Sheet1"], sep_cols)
    _enforce_separator_widths(wb["Sheet1"], sep_cols)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _csv_text_for_excel_open(value: object) -> str:
    """
    Excel's CSV importer treats long digit fields as numbers (scientific notation).

    A text formula like ``="199230130516"`` forces the full value as text in Excel.
    """
    s = plain_id_string(value)
    if s == "":
        return ""
    escaped = s.replace('"', '""')
    return f'="{escaped}"'


def dataframe_to_bytes(df: pd.DataFrame, fmt: ExportFormat) -> tuple[bytes, str, str]:
    """
    Serialize a DataFrame for Streamlit download_button.

    Returns
    -------
    data : bytes
    mime_type : str
    default_filename : str
    """
    export_df = coerce_dataframe_long_ids(_export_frame(df))

    if fmt == ExportFormat.CSV:
        csv_df = export_df.copy()
        for col in LONG_NUMERIC_ID_COLUMNS:
            if col in csv_df.columns:
                csv_df[col] = csv_df[col].map(_csv_text_for_excel_open)
        buffer = BytesIO()
        csv_df.to_csv(buffer, index=False, encoding=CSV_ENCODING)
        return buffer.getvalue(), "text/csv; charset=utf-8", "export.csv"

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Sheet1")

    xlsx_bytes = _finalize_xlsx_bytes(buffer)

    return xlsx_bytes, (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ), "export.xlsx"
